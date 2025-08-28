#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
üéØ –ì–∏–±—Ä–∏–¥–Ω—ã–π OCR —Ç–µ—Å—Ç–µ—Ä v13 (Google Cloud Vision Edition) - "–ë—Ä–æ–Ω–µ–±–æ–π–Ω–∞—è" –≤–µ—Ä—Å–∏—è
- –î–ª—è —ç–∫—Å—Ç—Ä–µ–º–∞–ª—å–Ω–æ –±–æ–ª—å—à–∏—Ö –∏–ª–∏ –Ω–µ—Å—Ç–∞–Ω–¥–∞—Ä—Ç–Ω—ã—Ö —Å—Ç—Ä–∞–Ω–∏—Ü PDF –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –Ω–∞–¥–µ–∂–Ω—ã–π –º–µ—Ç–æ–¥
  –∫–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏–∏ —á–µ—Ä–µ–∑ –≤—Ä–µ–º–µ–Ω–Ω—ã–µ —Ñ–∞–π–ª—ã –Ω–∞ –¥–∏—Å–∫–µ.
"""

import json
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple
import time
from datetime import datetime
import subprocess
import shutil
import io
import logging
import traceback
from logging.handlers import RotatingFileHandler

from PIL import Image
from google.api_core import exceptions as google_exceptions
from .file_utils import normalize_filename

Image.MAX_IMAGE_PIXELS = None

# ... (–æ—Å—Ç–∞–ª—å–Ω—ã–µ –∏–º–ø–æ—Ä—Ç—ã –±–µ–∑ –∏–∑–º–µ–Ω–µ–Ω–∏–π) ...
try:
    from google.cloud import vision
    GOOGLE_VISION_AVAILABLE = True
except ImportError:
    GOOGLE_VISION_AVAILABLE = False
try:
    import fitz
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False
try:
    from docx import Document as DocxDocument
    PYTHON_DOCX_AVAILABLE = True
except ImportError:
    PYTHON_DOCX_AVAILABLE = False
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

class OCRProcessor:
    # ... (init, _show_capabilities, get_available_dates, get_files_for_date, run_google_vision_ocr - –±–µ–∑ –∏–∑–º–µ–Ω–µ–Ω–∏–π) ...
    def __init__(self):
        self.data_dir = Path("data")
        self.attachments_dir = self.data_dir / "attachments"
        self.base_results_dir = self.data_dir / "final_results"
        self.texts_dir = self.base_results_dir / "texts"
        self.reports_dir = self.base_results_dir / "reports"
        self.logs_dir = self.data_dir / "logs" / "ocr"
        self.texts_dir.mkdir(parents=True, exist_ok=True)
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        self.logs_dir.mkdir(parents=True, exist_ok=True)
        
        # –ù–∞—Å—Ç—Ä–æ–π–∫–∞ –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è
        self._setup_logging()
        
        self.vision_client = vision.ImageAnnotatorClient() if GOOGLE_VISION_AVAILABLE else None
        
        if GOOGLE_VISION_AVAILABLE and self.vision_client:
            self.logger.info("Google Cloud Vision API —É—Å–ø–µ—à–Ω–æ –∏–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω")
        else:
            self.logger.error("Google Cloud Vision API –Ω–µ –∏–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω")
        
        self._show_capabilities()
        print("\n" + "=" * 70)
        print("üéØ OCR –¢–ï–°–¢–ï–† –° GOOGLE CLOUD VISION v13 üéØ")
        print(f"üìÅ –ò—Å—Ö–æ–¥–Ω—ã–µ —Ñ–∞–π–ª—ã: {self.attachments_dir}")
        print(f"üóÇÔ∏è  –†–µ–∑—É–ª—å—Ç–∞—Ç—ã –≤ –ø–∞–ø–∫–µ: {self.base_results_dir}")
        print("=" * 70)
    def _setup_logging(self):
        """üîß –ù–∞—Å—Ç—Ä–æ–π–∫–∞ —Å–∏—Å—Ç–µ–º—ã –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è —Å —Ä–æ—Ç–∞—Ü–∏–µ–π —Ñ–∞–π–ª–æ–≤"""
        
        # –°–æ–∑–¥–∞–µ–º –ª–æ–≥–≥–µ—Ä
        self.logger = logging.getLogger('OCRProcessor')
        self.logger.setLevel(logging.INFO)
        
        # –û—á–∏—â–∞–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–µ –æ–±—Ä–∞–±–æ—Ç—á–∏–∫–∏
        self.logger.handlers.clear()
        
        # –§–æ—Ä–º–∞—Ç—Ç–µ—Ä –¥–ª—è –ª–æ–≥–æ–≤
        formatter = logging.Formatter(
            '%(asctime)s | %(levelname)-8s | %(funcName)-20s | %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        
        # –§–∞–π–ª–æ–≤—ã–π –æ–±—Ä–∞–±–æ—Ç—á–∏–∫ —Å —Ä–æ—Ç–∞—Ü–∏–µ–π (–º–∞–∫—Å–∏–º—É–º 10MB, 5 —Ñ–∞–π–ª–æ–≤)
        log_file = self.logs_dir / 'ocr_processor.log'
        file_handler = RotatingFileHandler(
            log_file, 
            maxBytes=10*1024*1024,  # 10MB
            backupCount=5,
            encoding='utf-8'
        )
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)
        
        # –ö–æ–Ω—Å–æ–ª—å–Ω—ã–π –æ–±—Ä–∞–±–æ—Ç—á–∏–∫ –¥–ª—è –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∏—Ö –æ—à–∏–±–æ–∫
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.ERROR)
        console_handler.setFormatter(formatter)
        
        # –î–æ–±–∞–≤–ª—è–µ–º –æ–±—Ä–∞–±–æ—Ç—á–∏–∫–∏
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        
        self.logger.info("=" * 50)
        self.logger.info("OCR Processor –∑–∞–ø—É—â–µ–Ω")
        self.logger.info(f"–õ–æ–≥–∏ —Å–æ—Ö—Ä–∞–Ω—è—é—Ç—Å—è –≤: {log_file}")
        
    def _show_capabilities(self):
        antiword_ok = shutil.which('antiword') is not None
        print("üìã –í–û–ó–ú–û–ñ–ù–û–°–¢–ò –°–ò–°–¢–ï–ú–´:")
        if GOOGLE_VISION_AVAILABLE and self.vision_client:
            print("   ‚òÅÔ∏è Google Cloud Vision: ‚úÖ –ì–æ—Ç–æ–≤ –∫ —Ä–∞–±–æ—Ç–µ!")
        else:
            print("   ‚òÅÔ∏è Google Cloud Vision: ‚ùå –ù–ï –ù–ê–°–¢–†–û–ï–ù!")
        local_status = [f"PDF (—Ç–µ–∫—Å—Ç) {'‚úÖ' if PYMUPDF_AVAILABLE else '‚ùå'}", f"DOCX {'‚úÖ' if PYTHON_DOCX_AVAILABLE else '‚ùå'}", f"XLSX {'‚úÖ' if OPENPYXL_AVAILABLE else '‚ùå'}", f"DOC (antiword) {'‚úÖ' if antiword_ok else '‚ùå (brew install antiword)'}", f"XLS (xlrd) {'‚úÖ' if XLRD_AVAILABLE else '‚ùå'}"]
        print(f"   üìÑ –õ–æ–∫–∞–ª—å–Ω—ã–µ —Ñ–æ—Ä–º–∞—Ç—ã: {' | '.join(local_status)}")
    def get_available_dates(self) -> List[str]:
        if not self.attachments_dir.exists(): return []
        return sorted([d.name for d in self.attachments_dir.iterdir() if d.is_dir() and d.name.count("-") == 2])
    def get_files_for_date(self, date: str) -> List[Path]:
        date_dir = self.attachments_dir / date
        if not date_dir.exists():
            print(f"‚ùå –ü–∞–ø–∫–∞ –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç: {date_dir}")
            return []
        file_types = ["*.png", "*.jpg", "*.jpeg", "*.tiff", "*.pdf", "*.docx", "*.doc", "*.xlsx", "*.xls"]
        files = sorted(list(set(f for pat in file_types for f in date_dir.rglob(pat))))
        return files
    
    def _normalize_filename(self, filename: str) -> str:
        """üîß –ù–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏—è –∏–º–µ–Ω–∏ —Ñ–∞–π–ª–∞ —á–µ—Ä–µ–∑ –µ–¥–∏–Ω—É—é —Ñ—É–Ω–∫—Ü–∏—é –∏–∑ file_utils"""
        return normalize_filename(filename, remove_extension=True, to_lowercase=True)
    def run_google_vision_ocr(self, content: bytes) -> Tuple[str, float]:
        if not self.vision_client: raise RuntimeError("–ö–ª–∏–µ–Ω—Ç Google Vision –Ω–µ –∏–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω.")
        print("   ‚òÅÔ∏è –û—Ç–ø—Ä–∞–≤–∫–∞ –≤ Google Cloud Vision... (–º–æ–∂–µ—Ç –∑–∞–Ω—è—Ç—å –Ω–µ—Å–∫–æ–ª—å–∫–æ —Å–µ–∫—É–Ω–¥)")
        ts = time.time()
        image = vision.Image(content=content)
        response = self.vision_client.document_text_detection(image=image)
        if response.error.message:
            raise Exception(f"Google Vision API Error: {response.error.message}")
        text = response.full_text_annotation.text
        confidences = [page.confidence for page in response.full_text_annotation.pages]
        avg_confidence = np.mean(confidences) if confidences else 0.0
        elapsed = time.time() - ts
        print(f"   ‚ú® –ü–æ–ª—É—á–µ–Ω –æ—Ç–≤–µ—Ç –æ—Ç Google –∑–∞ {elapsed:.2f} —Å–µ–∫. –£–≤–µ—Ä–µ–Ω–Ω–æ—Å—Ç—å: {avg_confidence:.2%}")
        
        # –õ–æ–≥–∏—Ä—É–µ–º –≤—Ä–µ–º—è –≤—ã–ø–æ–ª–Ω–µ–Ω–∏—è –æ–ø–µ—Ä–∞—Ü–∏–∏
        self._log_operation_time("google_vision_ocr", ts, 
                               text_length=len(text) if text else 0, 
                               confidence=avg_confidence,
                               content_size_mb=len(content) / (1024 * 1024))
        
        return text, avg_confidence
    
    def run_google_vision_ocr_with_smart_compression(self, content: bytes, max_size_mb: float = 19.0) -> Tuple[str, float]:
        """
        –û—Ç–ø—Ä–∞–≤–∫–∞ –≤ Google Vision —Å –∏–Ω—Ç–µ–ª–ª–µ–∫—Ç—É–∞–ª—å–Ω—ã–º —Å–∂–∞—Ç–∏–µ–º
        """
        start_time = time.time()
        if not self.vision_client:
            raise RuntimeError("–ö–ª–∏–µ–Ω—Ç Google Vision –Ω–µ –∏–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω.")

        max_size_bytes = int(max_size_mb * 1024 * 1024)

        # –ï—Å–ª–∏ —Ä–∞–∑–º–µ—Ä –ø—Ä–∏–µ–º–ª–µ–º—ã–π, –æ—Ç–ø—Ä–∞–≤–ª—è–µ–º –∫–∞–∫ –µ—Å—Ç—å
        if len(content) <= max_size_bytes:
            result = self.run_google_vision_ocr(content)
            self._log_operation_time("google_vision_smart_compression", start_time, 
                                   compression_applied=False, 
                                   original_size_mb=len(content) / (1024 * 1024))
            return result

        print(f"   ‚ö†Ô∏è –†–∞–∑–º–µ—Ä –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏—è {len(content)/1024/1024:.1f}MB –ø—Ä–µ–≤—ã—à–∞–µ—Ç –ª–∏–º–∏—Ç {max_size_mb}MB")
        print("   üîß –ü—Ä–∏–º–µ–Ω—è—é –∏–Ω—Ç–µ–ª–ª–µ–∫—Ç—É–∞–ª—å–Ω–æ–µ —Å–∂–∞—Ç–∏–µ...")

        # –ó–∞–≥—Ä—É–∂–∞–µ–º –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ —á–µ—Ä–µ–∑ Pillow
        with Image.open(io.BytesIO(content)) as img:
            # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –≤ RGB –µ—Å–ª–∏ –Ω—É–∂–Ω–æ
            if img.mode not in ['RGB', 'L']:
                print(f"   üé® –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É—é –∏–∑ {img.mode} –≤ RGB")
                img = img.convert('RGB')
            
            original_size = img.size
            print(f"   üìè –ò—Å—Ö–æ–¥–Ω—ã–π —Ä–∞–∑–º–µ—Ä: {original_size[0]}x{original_size[1]} –ø–∏–∫—Å–µ–ª–µ–π")
            
            # –î–ª—è OCR –æ–ø—Ç–∏–º–∞–ª—å–Ω–æ–µ —Ä–∞–∑—Ä–µ—à–µ–Ω–∏–µ
            max_dimension = 2048  # –ú–∞–∫—Å–∏–º–∞–ª—å–Ω–∞—è —Å—Ç–æ—Ä–æ–Ω–∞
            
            # –í—ã—á–∏—Å–ª—è–µ–º –Ω–æ–≤—ã–π —Ä–∞–∑–º–µ—Ä —Å —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ–º –ø—Ä–æ–ø–æ—Ä—Ü–∏–π
            ratio = min(max_dimension / original_size[0], max_dimension / original_size[1])
            if ratio < 1:
                new_size = (int(original_size[0] * ratio), int(original_size[1] * ratio))
                print(f"   üìê –ò–∑–º–µ–Ω—è—é —Ä–∞–∑–º–µ—Ä –¥–æ: {new_size[0]}x{new_size[1]} –ø–∏–∫—Å–µ–ª–µ–π")
                img = img.resize(new_size, Image.Resampling.LANCZOS)
            
            # –ü—Ä–æ–±—É–µ–º —Ä–∞–∑–Ω—ã–µ —É—Ä–æ–≤–Ω–∏ –∫–∞—á–µ—Å—Ç–≤–∞ JPEG
            for quality in [95, 90, 85, 80, 75, 70]:
                buffer = io.BytesIO()
                
                # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤ JPEG —Å —Ç–µ–∫—É—â–∏–º –∫–∞—á–µ—Å—Ç–≤–æ–º
                img.save(buffer, format='JPEG', quality=quality, optimize=True)
                compressed_content = buffer.getvalue()
                
                size_mb = len(compressed_content) / 1024 / 1024
                print(f"   üéöÔ∏è –ö–∞—á–µ—Å—Ç–≤–æ {quality}%: {size_mb:.1f}MB")
                
                if len(compressed_content) <= max_size_bytes:
                        print(f"   ‚úÖ –ù–∞–π–¥–µ–Ω –ø–æ–¥—Ö–æ–¥—è—â–∏–π —Ä–∞–∑–º–µ—Ä: {size_mb:.1f}MB –ø—Ä–∏ –∫–∞—á–µ—Å—Ç–≤–µ {quality}%")
                        result = self.run_google_vision_ocr(compressed_content)
                        self._log_operation_time("google_vision_smart_compression", start_time, 
                                               compression_applied=True, 
                                               original_size_mb=len(content) / (1024 * 1024),
                                               final_size_mb=size_mb,
                                               quality_used=quality)
                        return result
            
            # –ï—Å–ª–∏ –¥–∞–∂–µ –ø—Ä–∏ 70% –∫–∞—á–µ—Å—Ç–≤–∞ —Ä–∞–∑–º–µ—Ä –±–æ–ª—å—à–æ–π, —É–º–µ–Ω—å—à–∞–µ–º —Ä–∞–∑—Ä–µ—à–µ–Ω–∏–µ –µ—â–µ –±–æ–ª—å—à–µ
            print("   ‚ö†Ô∏è –¢—Ä–µ–±—É–µ—Ç—Å—è –¥–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω–æ–µ —É–º–µ–Ω—å—à–µ–Ω–∏–µ —Ä–∞–∑—Ä–µ—à–µ–Ω–∏—è...")
            
            for max_dim in [1600, 1200, 1024, 800]:
                ratio = min(max_dim / img.size[0], max_dim / img.size[1])
                if ratio < 1:
                    new_size = (int(img.size[0] * ratio), int(img.size[1] * ratio))
                    resized_img = img.resize(new_size, Image.Resampling.LANCZOS)
                    
                    buffer = io.BytesIO()
                    resized_img.save(buffer, format='JPEG', quality=85, optimize=True)
                    compressed_content = buffer.getvalue()
                    
                    size_mb = len(compressed_content) / 1024 / 1024
                    print(f"   üìè –†–∞–∑–º–µ—Ä {new_size[0]}x{new_size[1]}: {size_mb:.1f}MB")
                    
                    if len(compressed_content) <= max_size_bytes:
                        print(f"   ‚úÖ –£—Å–ø–µ—à–Ω–æ–µ —Å–∂–∞—Ç–∏–µ –¥–æ {size_mb:.1f}MB")
                        result = self.run_google_vision_ocr(compressed_content)
                        self._log_operation_time("google_vision_smart_compression", start_time, 
                                               compression_applied=True, 
                                               original_size_mb=len(content) / (1024 * 1024),
                                               final_size_mb=size_mb,
                                               resolution_reduced=True,
                                               final_resolution=new_size)
                        return result
            
            self._log_operation_time("google_vision_smart_compression", start_time, 
                                   compression_applied=True, 
                                   original_size_mb=len(content) / (1024 * 1024),
                                   error="Failed to compress to acceptable size")
            raise RuntimeError("–ù–µ —É–¥–∞–ª–æ—Å—å —Å–∂–∞—Ç—å –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ –¥–æ –ø—Ä–∏–µ–º–ª–µ–º–æ–≥–æ —Ä–∞–∑–º–µ—Ä–∞")

    def extract_text_from_file(self, file_path: Path, date: str = None) -> Dict:
        """üîç –ò–∑–≤–ª–µ—á–µ–Ω–∏–µ —Ç–µ–∫—Å—Ç–∞ –∏–∑ —Ñ–∞–π–ª–∞ —Å –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏–º –≤—ã–±–æ—Ä–æ–º –º–µ—Ç–æ–¥–∞"""
        
        start_time = time.time()
        file_name = file_path.name
        file_size_mb = file_path.stat().st_size / (1024 * 1024)
        
        # –õ–æ–≥–∏—Ä—É–µ–º –Ω–∞—á–∞–ª–æ –æ–±—Ä–∞–±–æ—Ç–∫–∏
        self.logger.info(f"–ù–∞—á–∞–ª–æ –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Ñ–∞–π–ª–∞: {file_name} ({file_size_mb:.2f} MB) –¥–ª—è –¥–∞—Ç—ã {date}")
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ —É–∂–µ –æ–±—Ä–∞–±–æ—Ç–∞–Ω–Ω—ã–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã
        if date and self._check_existing_results(file_path, date):
            self.logger.info(f"–§–∞–π–ª {file_name} —É–∂–µ –æ–±—Ä–∞–±–æ—Ç–∞–Ω, –∏—Å–ø–æ–ª—å–∑—É–µ–º –∫—ç—à–∏—Ä–æ–≤–∞–Ω–Ω—ã–π —Ä–µ–∑—É–ª—å—Ç–∞—Ç")
            print(f"   ‚è≠Ô∏è –í–ª–æ–∂–µ–Ω–∏–µ {file_name} —É–∂–µ –æ–±—Ä–∞–±–æ—Ç–∞–Ω–æ. –ü—Ä–æ–ø—É—Å–∫–∞—é.")
            return self._get_existing_result(file_path, date)
        
        self.logger.info(f"–û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–∞ {file_name} –º–µ—Ç–æ–¥–æ–º OCR")
        print(f"   üîÑ –û–±—Ä–∞–±–∞—Ç—ã–≤–∞—é –≤–ª–æ–∂–µ–Ω–∏–µ {file_name} ({file_size_mb:.1f} MB)")
        
        result = {
            "file_name": file_name,
            "file_path": str(file_path),
            "file_size_mb": round(file_size_mb, 2),
            "success": False,
            "text": "",
            "method": "unknown",
            "confidence": 0.0,
            "processing_time_sec": 0.0,
            "error": None
        }
        
        ext = file_path.suffix.lower()
        text, method, confidence, error = "", "unknown", 0.0, None
        ts = time.time()
        try:
            if ext == ".docx":
                print("   üìÑ –û–±—Ä–∞–±–æ—Ç–∫–∞ DOCX –ª–æ–∫–∞–ª—å–Ω–æ...")
                doc = DocxDocument(file_path)
                text = "\n".join([p.text for p in doc.paragraphs])
                method, confidence = "local_docx", 1.0
            elif ext == ".doc":
                print("   üìÑ –û–±—Ä–∞–±–æ—Ç–∫–∞ DOC (—Å—Ç–∞—Ä—ã–π —Ñ–æ—Ä–º–∞—Ç) —á–µ—Ä–µ–∑ antiword...")
                if not shutil.which('antiword'):
                    raise FileNotFoundError("–£—Ç–∏–ª–∏—Ç–∞ 'antiword' –Ω–µ –Ω–∞–π–¥–µ–Ω–∞. –£—Å—Ç–∞–Ω–æ–≤–∏—Ç–µ –µ–µ: brew install antiword")
                process = subprocess.run(['antiword', str(file_path)], capture_output=True, text=True, encoding='utf-8', errors='ignore')
                if process.returncode == 0:
                    text = process.stdout
                else:
                    raise RuntimeError(f"Antiword –≤–µ—Ä–Ω—É–ª –æ—à–∏–±–∫—É: {process.stderr}")
                method, confidence = "local_doc_antiword", 1.0
            elif ext == ".xlsx":
                print("   üìÑ –û–±—Ä–∞–±–æ—Ç–∫–∞ XLSX –ª–æ–∫–∞–ª—å–Ω–æ...")
                wb = openpyxl.load_workbook(file_path, data_only=True)
                lines = [" | ".join([str(cell.value or "") for cell in row]) for sheet in wb.worksheets for row in sheet.iter_rows()]
                text = "\n".join(lines)
                method, confidence = "local_xlsx", 1.0
            elif ext == ".xls":
                print("   üìÑ –û–±—Ä–∞–±–æ—Ç–∫–∞ XLS (—Å—Ç–∞—Ä—ã–π —Ñ–æ—Ä–º–∞—Ç) –ª–æ–∫–∞–ª—å–Ω–æ...")
                if not XLRD_AVAILABLE: raise ImportError("–ë–∏–±–ª–∏–æ—Ç–µ–∫–∞ xlrd –Ω–µ –Ω–∞–π–¥–µ–Ω–∞. –£—Å—Ç–∞–Ω–æ–≤–∏—Ç–µ: pip install xlrd")
                wb = xlrd.open_workbook(file_path, encoding_override="cp1251")
                lines = []
                for sheet in wb.sheets():
                    for row_idx in range(sheet.nrows):
                        lines.append(" | ".join([str(sheet.cell(row_idx, col_idx).value or "") for col_idx in range(sheet.ncols)]))
                text = "\n".join(lines)
                method, confidence = "local_xls", 1.0

            # <<< –ò–ó–ú–ï–ù–ï–ù–ò–ï: –°–∞–º–∞—è –Ω–∞–¥–µ–∂–Ω–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ PDF >>>
            elif ext == ".pdf":
                print("   üìÑ –û–±—Ä–∞–±–æ—Ç–∫–∞ PDF... –ü–æ–ø—ã—Ç–∫–∞ –∏–∑–≤–ª–µ—á—å —Ç–µ–∫—Å—Ç–æ–≤—ã–π —Å–ª–æ–π.")
                doc = fitz.open(file_path)
                texts = [page.get_text() for page in doc]
                full_text_direct = "\n\n".join(texts).strip()
                
                if len(full_text_direct) > 100:
                    print("   ‚úÖ –û–±–Ω–∞—Ä—É–∂–µ–Ω —Ç–µ–∫—Å—Ç–æ–≤—ã–π —Å–ª–æ–π. –ò–∑–≤–ª–µ—á–µ–Ω–æ –ª–æ–∫–∞–ª—å–Ω–æ.")
                    text, method, confidence = full_text_direct, "local_pdf_text", 1.0
                else:
                    print("   üñºÔ∏è –¢–µ–∫—Å—Ç–æ–≤—ã–π —Å–ª–æ–π –ø—É—Å—Ç. –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º —Å—Ç—Ä–∞–Ω–∏—Ü—ã PDF –≤ –∫–∞—Ä—Ç–∏–Ω–∫–∏ –¥–ª—è Google Vision.")
                    all_pages_text = []
                    all_confidences = []
                    
                    for page_idx, page in enumerate(doc):
                        print(f"     -- –û–±—Ä–∞–±–æ—Ç–∫–∞ —Å—Ç—Ä–∞–Ω–∏—Ü—ã {page_idx+1}/{len(doc)} --")
                        
                        # –ò—Å–ø–æ–ª—å–∑—É–µ–º –æ–ø—Ç–∏–º–∞–ª—å–Ω–æ–µ DPI –¥–ª—è OCR
                        dpi = 200  # –î–æ—Å—Ç–∞—Ç–æ—á–Ω–æ –¥–ª—è –∫–∞—á–µ—Å—Ç–≤–µ–Ω–Ω–æ–≥–æ OCR
                        
                        try:
                            pix = page.get_pixmap(dpi=dpi)
                            img_bytes = pix.tobytes("png")
                            
                            # –ò—Å–ø–æ–ª—å–∑—É–µ–º –Ω–æ–≤—ã–π –º–µ—Ç–æ–¥ —Å –∏–Ω—Ç–µ–ª–ª–µ–∫—Ç—É–∞–ª—å–Ω—ã–º —Å–∂–∞—Ç–∏–µ–º
                            page_text, page_confidence = self.run_google_vision_ocr_with_smart_compression(img_bytes)
                            
                        except google_exceptions.InvalidArgument as e:
                            print(f"     ‚ùå –û—à–∏–±–∫–∞ Google Vision: {str(e)[:100]}...")
                            
                            # –ü—Ä–æ–±—É–µ–º —Å –µ—â–µ –º–µ–Ω—å—à–∏–º DPI
                            print("     üîß –ü—Ä–æ–±—É—é —Å —É–º–µ–Ω—å—à–µ–Ω–Ω—ã–º —Ä–∞–∑—Ä–µ—à–µ–Ω–∏–µ–º (150 DPI)...")
                            try:
                                pix_low = page.get_pixmap(dpi=150)
                                img_bytes_low = pix_low.tobytes("png")
                                page_text, page_confidence = self.run_google_vision_ocr_with_smart_compression(img_bytes_low)
                            except Exception as e2:
                                print(f"     ‚ùå –ö—Ä–∏—Ç–∏—á–µ—Å–∫–∞—è –æ—à–∏–±–∫–∞: {e2}")
                                page_text, page_confidence = f"[–û–®–ò–ë–ö–ê –û–ë–†–ê–ë–û–¢–ö–ò –°–¢–†–ê–ù–ò–¶–´]", 0.0
                        
                        all_pages_text.append(page_text)
                        all_confidences.append(page_confidence)
                    
                    text = "\n\n--- PAGE BREAK ---\n\n".join(all_pages_text)
                    confidence = np.mean(all_confidences) if all_confidences else 0.0
                    method = "google_vision_pdf_optimized"
            
            elif ext in [".png", ".jpg", ".jpeg", ".tiff"]:
                print(f"   üñºÔ∏è –û–±—Ä–∞–±–æ—Ç–∫–∞ –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏—è ({ext}). –û—Ç–ø—Ä–∞–≤–ª—è–µ–º –≤ Google Vision.")
                try:
                    img_bytes = file_path.read_bytes()
                    text, confidence = self.run_google_vision_ocr_with_smart_compression(img_bytes)
                except Exception as e:
                    print(f"     ‚ùå –û—à–∏–±–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏—è: {e}")
                    text, confidence = f"[–û–®–ò–ë–ö–ê –û–ë–†–ê–ë–û–¢–ö–ò –ò–ó–û–ë–†–ê–ñ–ï–ù–ò–Ø]", 0.0
                method = "google_vision_image_optimized"

            else:
                method, error = "unsupported", f"–§–æ—Ä–º–∞—Ç {ext} –Ω–µ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ—Ç—Å—è."

        except Exception as e:
            error = str(e)
            method = "error"
            error_details = {
                "file_name": file_name,
                "file_size_mb": file_size_mb,
                "error_type": type(e).__name__,
                "error_message": str(e),
                "traceback": traceback.format_exc()
            }
            self.logger.error(f"–û—à–∏–±–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Ñ–∞–π–ª–∞ {file_name}: {e}", extra=error_details)
            print(f"   ‚ùå –ü—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞: {e}")

        # –§–æ—Ä–º–∏—Ä—É–µ–º —Ä–µ–∑—É–ª—å—Ç–∞—Ç
        processing_time = time.time() - ts
        total_time = time.time() - start_time
        
        result.update({
            "success": not error and bool(text.strip()),
            "text": text.strip(),
            "method": method,
            "confidence": confidence,
            "error": error,
            "processing_time_sec": processing_time,
            "total_time_sec": total_time,
            "timestamp": datetime.now().isoformat()
        })
        
        # –õ–æ–≥–∏—Ä—É–µ–º —Ä–µ–∑—É–ª—å—Ç–∞—Ç –æ–±—Ä–∞–±–æ—Ç–∫–∏
        if result["success"]:
            self.logger.info(f"–£—Å–ø–µ—à–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∞–Ω {file_name}: –º–µ—Ç–æ–¥={method}, –≤—Ä–µ–º—è={total_time:.2f}—Å, —É–≤–µ—Ä–µ–Ω–Ω–æ—Å—Ç—å={confidence:.2%}")
        else:
            self.logger.warning(f"–ù–µ—É–¥–∞—á–Ω–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ {file_name}: –º–µ—Ç–æ–¥={method}, –æ—à–∏–±–∫–∞={error}")
        
        # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ä–µ–∑—É–ª—å—Ç–∞—Ç, –µ—Å–ª–∏ —É–∫–∞–∑–∞–Ω–∞ –¥–∞—Ç–∞
        if date:
            self.save_result(result, date)
        
        return result
    
    def _check_existing_results(self, file_path: Path, date: str) -> bool:
        """üîç –ü—Ä–æ–≤–µ—Ä–∫–∞ —Å—É—â–µ—Å—Ç–≤–æ–≤–∞–Ω–∏—è —É–∂–µ –æ–±—Ä–∞–±–æ—Ç–∞–Ω–Ω—ã—Ö —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤"""
        
        date_texts_dir = self.texts_dir / date
        if not date_texts_dir.exists():
            self.logger.debug(f"–ü–∞–ø–∫–∞ –¥–ª—è –¥–∞—Ç—ã {date} –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç: {date_texts_dir}")
            return False
        
        # –ü–æ–ª—É—á–∞–µ–º —Ç–æ—á–Ω–æ–µ –∏–º—è —Ñ–∞–π–ª–∞ –±–µ–∑ —Ä–∞—Å—à–∏—Ä–µ–Ω–∏—è
        file_stem = file_path.stem
        
        # –ò—â–µ–º —Ñ–∞–π–ª—ã —Å —Ç–æ—á–Ω—ã–º —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ–º –∏–º–µ–Ω–∏ (–≤–∫–ª—é—á–∞—è —Ñ–∞–π–ª—ã-–º–∞—Ä–∫–µ—Ä—ã –æ—à–∏–±–æ–∫)
        existing_files = list(date_texts_dir.glob(f"{file_stem}___*.txt"))
        
        # –ï—Å–ª–∏ –Ω–∞–π–¥–µ–Ω—ã —Ñ–∞–π–ª—ã —Å —Ç–æ—á–Ω—ã–º —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ–º, –≤–æ–∑–≤—Ä–∞—â–∞–µ–º True
        if existing_files:
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ —Å—Ä–µ–¥–∏ –Ω–∞–π–¥–µ–Ω–Ω—ã—Ö —Ñ–∞–π–ª–æ–≤ —Ñ–∞–π–ª—ã-–º–∞—Ä–∫–µ—Ä—ã –æ—à–∏–±–æ–∫
            error_files = [f for f in existing_files if '_ERROR.txt' in f.name]
            success_files = [f for f in existing_files if '_ERROR.txt' not in f.name]
            
            if success_files:
                self.logger.debug(f"–ù–∞–π–¥–µ–Ω—ã —É—Å–ø–µ—à–Ω—ã–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã –¥–ª—è {file_path.name}: {len(success_files)} —Ñ–∞–π–ª–æ–≤")
                return True
            elif error_files:
                self.logger.debug(f"–ù–∞–π–¥–µ–Ω—ã —Ç–æ–ª—å–∫–æ —Ñ–∞–π–ª—ã-–º–∞—Ä–∫–µ—Ä—ã –æ—à–∏–±–æ–∫ –¥–ª—è {file_path.name}: {len(error_files)} —Ñ–∞–π–ª–æ–≤")
                # –í–æ–∑–≤—Ä–∞—â–∞–µ–º True, —á—Ç–æ–±—ã –Ω–µ –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞—Ç—å –ø–æ–≤—Ç–æ—Ä–Ω–æ —Ñ–∞–π–ª—ã —Å –æ—à–∏–±–∫–∞–º–∏
                return True
        
        # –î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω–æ –ø—Ä–æ–≤–µ—Ä—è–µ–º –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω–æ–µ –∏–º—è –¥–ª—è —Å–æ–≤–º–µ—Å—Ç–∏–º–æ—Å—Ç–∏ —Å–æ —Å—Ç–∞—Ä—ã–º–∏ —Ñ–∞–π–ª–∞–º–∏
        normalized_name = self._normalize_filename(file_stem)
        normalized_files = list(date_texts_dir.glob(f"{normalized_name}___*.txt"))
        
        if normalized_files:
            # –ê–Ω–∞–ª–æ–≥–∏—á–Ω–æ –ø—Ä–æ–≤–µ—Ä—è–µ–º –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω—ã–µ —Ñ–∞–π–ª—ã
            error_files = [f for f in normalized_files if '_ERROR.txt' in f.name]
            success_files = [f for f in normalized_files if '_ERROR.txt' not in f.name]
            
            if success_files:
                self.logger.debug(f"–ù–∞–π–¥–µ–Ω—ã –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω—ã–µ —É—Å–ø–µ—à–Ω—ã–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã –¥–ª—è {file_path.name}: {len(success_files)} —Ñ–∞–π–ª–æ–≤")
                return True
            elif error_files:
                self.logger.debug(f"–ù–∞–π–¥–µ–Ω—ã –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω—ã–µ —Ñ–∞–π–ª—ã-–º–∞—Ä–∫–µ—Ä—ã –æ—à–∏–±–æ–∫ –¥–ª—è {file_path.name}: {len(error_files)} —Ñ–∞–π–ª–æ–≤")
                return True
        
        self.logger.debug(f"–†–µ–∑—É–ª—å—Ç–∞—Ç—ã –¥–ª—è {file_path.name} –Ω–µ –Ω–∞–π–¥–µ–Ω—ã")
        return False
    
    def _get_existing_result(self, file_path: Path, date: str) -> Dict:
        """üìÑ –ü–æ–ª—É—á–µ–Ω–∏–µ —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–≥–æ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏"""
        
        date_texts_dir = self.texts_dir / date
        file_stem = file_path.stem
        
        # –°–Ω–∞—á–∞–ª–∞ –∏—â–µ–º —Ñ–∞–π–ª—ã —Å —Ç–æ—á–Ω—ã–º —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ–º –∏–º–µ–Ω–∏
        existing_files = list(date_texts_dir.glob(f"{file_stem}___*.txt"))
        
        # –ï—Å–ª–∏ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ, –∏—â–µ–º –ø–æ –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω–æ–º—É –∏–º–µ–Ω–∏
        if not existing_files:
            normalized_name = self._normalize_filename(file_stem)
            existing_files = list(date_texts_dir.glob(f"{normalized_name}___*.txt"))
        
        if not existing_files:
            # –ï—Å–ª–∏ —Ñ–∞–π–ª–æ–≤ –Ω–µ—Ç, –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –ø—É—Å—Ç–æ–π —Ä–µ–∑—É–ª—å—Ç–∞—Ç
            return {
                "file_name": file_path.name,
                "file_path": str(file_path),
                "file_size_mb": round(file_path.stat().st_size / (1024 * 1024), 2),
                "success": False,
                "text": "",
                "method": "not_found",
                "confidence": 0.0,
                "processing_time_sec": 0.0,
                "error": "Existing result not found"
            }
        
        # –†–∞–∑–¥–µ–ª—è–µ–º —Ñ–∞–π–ª—ã –Ω–∞ —É—Å–ø–µ—à–Ω—ã–µ –∏ —Ñ–∞–π–ª—ã-–º–∞—Ä–∫–µ—Ä—ã –æ—à–∏–±–æ–∫
        error_files = [f for f in existing_files if '_ERROR.txt' in f.name]
        success_files = [f for f in existing_files if '_ERROR.txt' not in f.name]
        
        # –ü—Ä–∏–æ—Ä–∏—Ç–µ—Ç –æ—Ç–¥–∞–µ–º —É—Å–ø–µ—à–Ω—ã–º —Ñ–∞–π–ª–∞–º
        if success_files:
            result_file = success_files[0]
            is_error = False
        elif error_files:
            result_file = error_files[0]
            is_error = True
        else:
            result_file = existing_files[0]
            is_error = '_ERROR.txt' in result_file.name
        
        method = result_file.stem.split('___')[-1] if '___' in result_file.stem else 'unknown'
        
        try:
            with open(result_file, 'r', encoding='utf-8') as f:
                content = f.read()
                
            # –ò–∑–≤–ª–µ–∫–∞–µ–º —Ç–µ–∫—Å—Ç (–ø—Ä–æ–ø—É—Å–∫–∞–µ–º –∑–∞–≥–æ–ª–æ–≤–∫–∏)
            lines = content.split('\n')
            text_start_idx = 0
            for i, line in enumerate(lines):
                if line.startswith('# ==='):
                    text_start_idx = i + 2
                    break
            
            text = '\n'.join(lines[text_start_idx:]).strip()
            
            if is_error:
                # –î–ª—è —Ñ–∞–π–ª–æ–≤-–º–∞—Ä–∫–µ—Ä–æ–≤ –æ—à–∏–±–æ–∫ –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ–± –æ—à–∏–±–∫–µ
                return {
                    "file_name": file_path.name,
                    "file_path": str(file_path),
                    "file_size_mb": round(file_path.stat().st_size / (1024 * 1024), 2),
                    "success": False,
                    "text": "",
                    "method": f"{method}_cached",
                    "confidence": 0.0,
                    "processing_time_sec": 0.0,
                    "error": "Cached error result"
                }
            else:
                return {
                    "file_name": file_path.name,
                    "file_path": str(file_path),
                    "file_size_mb": round(file_path.stat().st_size / (1024 * 1024), 2),
                    "success": True,
                    "text": text,
                    "method": f"{method}_cached",
                    "confidence": 1.0,  # –ü—Ä–µ–¥–ø–æ–ª–∞–≥–∞–µ–º –≤—ã—Å–æ–∫—É—é —É–≤–µ—Ä–µ–Ω–Ω–æ—Å—Ç—å –¥–ª—è –∫—ç—à–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤
                    "processing_time_sec": 0.0,
                    "error": None
                }
            
        except Exception as e:
            return {
                "file_name": file_path.name,
                "file_path": str(file_path),
                "file_size_mb": round(file_path.stat().st_size / (1024 * 1024), 2),
                "success": False,
                "text": "",
                "method": "cache_error",
                "confidence": 0.0,
                "processing_time_sec": 0.0,
                "error": f"Error reading cached result: {e}"
            }

    # ... (–≤—Å–µ –æ—Å—Ç–∞–ª—å–Ω—ã–µ —Ñ—É–Ω–∫—Ü–∏–∏: save_result, _print_summary, test_files_by_date, main - –±–µ–∑ –∏–∑–º–µ–Ω–µ–Ω–∏–π) ...
    def _verify_saved_result(self, result: Dict, date: str) -> bool:
        """üîç –ü—Ä–æ–≤–µ—Ä–∫–∞ –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ—Å—Ç–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–Ω–æ–≥–æ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞"""
        try:
            file_name = result['file_name']
            txt_file_path = result.get('txt_file_path')
            
            if not txt_file_path:
                self.logger.error(f"–û—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç –ø—É—Ç—å –∫ TXT —Ñ–∞–π–ª—É –¥–ª—è {file_name}")
                return False
            
            txt_path = Path(txt_file_path)
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Å—É—â–µ—Å—Ç–≤–æ–≤–∞–Ω–∏–µ —Ñ–∞–π–ª–∞
            if not txt_path.exists():
                self.logger.error(f"TXT —Ñ–∞–π–ª –Ω–µ –Ω–∞–π–¥–µ–Ω: {txt_path}")
                return False
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ä–∞–∑–º–µ—Ä —Ñ–∞–π–ª–∞ (–¥–æ–ª–∂–µ–Ω –±—ã—Ç—å –±–æ–ª—å—à–µ 0)
            if txt_path.stat().st_size == 0:
                self.logger.error(f"TXT —Ñ–∞–π–ª –ø—É—Å—Ç–æ–π: {txt_path}")
                return False
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º –≤–æ–∑–º–æ–∂–Ω–æ—Å—Ç—å —á—Ç–µ–Ω–∏—è —Ñ–∞–π–ª–∞
            try:
                with open(txt_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ –æ–±—è–∑–∞—Ç–µ–ª—å–Ω—ã—Ö –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤
                if not content.startswith('# üìÑ –§–∞–π–ª:'):
                    self.logger.error(f"–ù–µ–≤–µ—Ä–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç TXT —Ñ–∞–π–ª–∞: {txt_path}")
                    return False
                    
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ —Ä–∞–∑–¥–µ–ª–∏—Ç–µ–ª—è
                if '# =' not in content:
                    self.logger.error(f"–û—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç —Ä–∞–∑–¥–µ–ª–∏—Ç–µ–ª—å –≤ TXT —Ñ–∞–π–ª–µ: {txt_path}")
                    return False
                    
            except Exception as e:
                self.logger.error(f"–û—à–∏–±–∫–∞ —á—Ç–µ–Ω–∏—è TXT —Ñ–∞–π–ª–∞ {txt_path}: {e}")
                return False
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º JSON –æ—Ç—á–µ—Ç
            json_report_path = self.reports_dir / f"test_report_{date}.json"
            if json_report_path.exists():
                try:
                    with open(json_report_path, 'r', encoding='utf-8') as f:
                        report_data = json.load(f)
                        
                    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ —Ä–µ–∑—É–ª—å—Ç–∞—Ç –µ—Å—Ç—å –≤ JSON –æ—Ç—á–µ—Ç–µ
                    found_in_report = False
                    for entry in report_data:
                        if entry.get('file_name') == file_name:
                            found_in_report = True
                            break
                    
                    if not found_in_report:
                        self.logger.warning(f"–†–µ–∑—É–ª—å—Ç–∞—Ç –¥–ª—è {file_name} –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ JSON –æ—Ç—á–µ—Ç–µ")
                        # –ù–µ —Å—á–∏—Ç–∞–µ–º —ç—Ç–æ –∫—Ä–∏—Ç–∏—á–µ—Å–∫–æ–π –æ—à–∏–±–∫–æ–π
                        
                except Exception as e:
                    self.logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–æ–≤–µ—Ä–∫–∏ JSON –æ—Ç—á–µ—Ç–∞ {json_report_path}: {e}")
                    return False
            
            self.logger.debug(f"–í–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏—è —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞ –¥–ª—è {file_name} –ø—Ä–æ—à–ª–∞ —É—Å–ø–µ—à–Ω–æ")
            return True
            
        except Exception as e:
            self.logger.error(f"–û—à–∏–±–∫–∞ –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏–∏ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞ –¥–ª—è {result.get('file_name', 'unknown')}: {e}")
            return False

    def save_result(self, result: Dict, date: str):
        """üíæ –°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ OCR –≤ —Ñ–∞–π–ª—ã —Å –ø—Ä–æ–≤–µ—Ä–∫–æ–π –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ—Å—Ç–∏"""
        date_texts_dir = self.texts_dir / date
        date_texts_dir.mkdir(parents=True, exist_ok=True)
        
        file_name = result['file_name']
        
        try:
            # –°–æ—Ö—Ä–∞–Ω—è–µ–º TXT —Ñ–∞–π–ª –í–°–ï–ì–î–ê –ø—Ä–∏ —É—Å–ø–µ—à–Ω–æ–π –æ–±—Ä–∞–±–æ—Ç–∫–µ (–¥–∞–∂–µ –µ—Å–ª–∏ —Ç–µ–∫—Å—Ç –ø—É—Å—Ç–æ–π)
            # –≠—Ç–æ –ø—Ä–µ–¥–æ—Ç–≤—Ä–∞—â–∞–µ—Ç –ø–æ–≤—Ç–æ—Ä–Ω—É—é –æ–±—Ä–∞–±–æ—Ç–∫—É —Ñ–∞–π–ª–æ–≤ —Å –ø—É—Å—Ç—ã–º —Å–æ–¥–µ—Ä–∂–∏–º—ã–º
            if result["success"]:
                # –ò—Å–ø–æ–ª—å–∑—É–µ–º –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω–æ–µ –∏–º—è —Ñ–∞–π–ª–∞ –¥–ª—è –ø—Ä–µ–¥–æ—Ç–≤—Ä–∞—â–µ–Ω–∏—è –¥—É–±–ª–∏—Ä–æ–≤–∞–Ω–∏—è
                normalized_name = self._normalize_filename(Path(result['file_name']).stem)
                txt_filename = f"{normalized_name}___{result['method']}.txt"
                txt_path = date_texts_dir / txt_filename
                
                # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Å–æ–¥–µ—Ä–∂–∏–º–æ–µ –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è
                text_content = result.get("text", "").strip()
                if not text_content:
                    text_content = "[–§–ê–ô–õ –û–ë–†–ê–ë–û–¢–ê–ù –£–°–ü–ï–®–ù–û, –ù–û –¢–ï–ö–°–¢ –ù–ï –ò–ó–í–õ–ï–ß–ï–ù]"
                
                with open(txt_path, "w", encoding="utf-8") as f:
                    f.write(f"# üìÑ –§–∞–π–ª: {result['file_name']}\n# ‚öôÔ∏è –ú–µ—Ç–æ–¥: {result['method']}\n")
                    f.write(f"# ‚ú® –£–≤–µ—Ä–µ–Ω–Ω–æ—Å—Ç—å: {result['confidence']:.2%}\n# ‚è±Ô∏è –í—Ä–µ–º—è: {result['processing_time_sec']:.2f} —Å–µ–∫\n")
                    if result.get('error'):
                        f.write(f"# ‚ö†Ô∏è –û—à–∏–±–∫–∞: {result['error']}\n")
                    f.write("# " + "=" * 50 + "\n\n" + text_content)
                
                result["txt_file_path"] = str(txt_path)
                self.logger.info(f"–°–æ—Ö—Ä–∞–Ω–µ–Ω TXT —Ñ–∞–π–ª –¥–ª—è {file_name}: {txt_path} (—Ç–µ–∫—Å—Ç: {len(text_content)} —Å–∏–º–≤–æ–ª–æ–≤)")
            else:
                # –î–ª—è –Ω–µ—É—Å–ø–µ—à–Ω–æ–π –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Ç–æ–∂–µ —Å–æ–∑–¥–∞–µ–º —Ñ–∞–π–ª-–º–∞—Ä–∫–µ—Ä
                normalized_name = self._normalize_filename(Path(result['file_name']).stem)
                txt_filename = f"{normalized_name}___{result['method']}_ERROR.txt"
                txt_path = date_texts_dir / txt_filename
                
                with open(txt_path, "w", encoding="utf-8") as f:
                    f.write(f"# üìÑ –§–∞–π–ª: {result['file_name']}\n# ‚öôÔ∏è –ú–µ—Ç–æ–¥: {result['method']}\n")
                    f.write(f"# ‚ùå –û–®–ò–ë–ö–ê –û–ë–†–ê–ë–û–¢–ö–ò\n# ‚ö†Ô∏è –û—à–∏–±–∫–∞: {result.get('error', '–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–∞—è –æ—à–∏–±–∫–∞')}\n")
                    f.write(f"# ‚è±Ô∏è –í—Ä–µ–º—è: {result['processing_time_sec']:.2f} —Å–µ–∫\n")
                    f.write("# " + "=" * 50 + "\n\n[–§–ê–ô–õ –ù–ï –£–î–ê–õ–û–°–¨ –û–ë–†–ê–ë–û–¢–ê–¢–¨]")
                
                result["txt_file_path"] = str(txt_path)
                self.logger.warning(f"–°–æ—Ö—Ä–∞–Ω–µ–Ω —Ñ–∞–π–ª-–º–∞—Ä–∫–µ—Ä –æ—à–∏–±–∫–∏ –¥–ª—è {file_name}: {txt_path}")
            
            # –í—Å–µ–≥–¥–∞ —Å–æ—Ö—Ä–∞–Ω—è–µ–º JSON –æ—Ç—á–µ—Ç (–¥–∞–∂–µ –ø—Ä–∏ –æ—à–∏–±–∫–∞—Ö)
            json_report_path = self.reports_dir / f"test_report_{date}.json"
            report_data = []
            
            if json_report_path.exists():
                with open(json_report_path, "r", encoding="utf-8") as f:
                    try: 
                        report_data = json.load(f)
                    except json.JSONDecodeError as e:
                        self.logger.error(f"–û—à–∏–±–∫–∞ —á—Ç–µ–Ω–∏—è JSON –æ—Ç—á–µ—Ç–∞ {json_report_path}: {e}")
                        report_data = []
            
            report_data.append(result)
            
            with open(json_report_path, "w", encoding="utf-8") as f:
                json.dump(report_data, f, ensure_ascii=False, indent=2)
            
            self.logger.info(f"–û–±–Ω–æ–≤–ª–µ–Ω JSON –æ—Ç—á–µ—Ç –¥–ª—è –¥–∞—Ç—ã {date}: {len(report_data)} –∑–∞–ø–∏—Å–µ–π")
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ—Å—Ç—å —Å–æ—Ö—Ä–∞–Ω–µ–Ω–Ω—ã—Ö —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤
            if self._verify_saved_result(result, date):
                print(f"   ‚úÖ –†–µ–∑—É–ª—å—Ç–∞—Ç—ã —Ä–∞—Å–ø–æ–∑–Ω–∞–≤–∞–Ω–∏—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω—ã –∏ –ø—Ä–æ–≤–µ—Ä–µ–Ω—ã")
            else:
                print(f"   ‚ö†Ô∏è –†–µ–∑—É–ª—å—Ç–∞—Ç—ã —Å–æ—Ö—Ä–∞–Ω–µ–Ω—ã, –Ω–æ –æ–±–Ω–∞—Ä—É–∂–µ–Ω—ã –ø—Ä–æ–±–ª–µ–º—ã –ø—Ä–∏ –ø—Ä–æ–≤–µ—Ä–∫–µ")
                self.logger.warning(f"–ü—Ä–æ–±–ª–µ–º—ã –ø—Ä–∏ –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏–∏ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞ –¥–ª—è {file_name}")
            
        except Exception as e:
            error_msg = f"–û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ –¥–ª—è {file_name}: {e}"
            self.logger.error(error_msg, exc_info=True)
            print(f"   ‚ùå {error_msg}")
    def _print_summary(self, date: str, stats: Dict):
        """üìä –í—ã–≤–æ–¥ –∏—Ç–æ–≥–æ–≤–æ–π —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏ –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Ñ–∞–π–ª–æ–≤ —Å –¥–µ—Ç–∞–ª–∏–∑–∞—Ü–∏–µ–π"""
        print("\n" + "="*25 + f" üìä –ò–¢–û–ì–ò –¢–ï–°–¢–ê –ó–ê {date} " + "="*25)
        
        total = stats.get('total', 0)
        successful = stats.get('successful', 0)
        skipped = stats.get('skipped', 0)
        processed = stats.get('processed', 0)
        failed = processed - successful
        
        success_rate = (successful / processed * 100) if processed > 0 else 0
        
        verification_passed = stats.get('verification_passed', 0)
        verification_failed = stats.get('verification_failed', 0)
        verification_rate = (verification_passed / successful * 100) if successful > 0 else 0
        
        # –û—Å–Ω–æ–≤–Ω–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞
        print(f"üóÇÔ∏è  –í—Å–µ–≥–æ —Ñ–∞–π–ª–æ–≤ –ø—Ä–æ–≤–µ—Ä–µ–Ω–æ: {total}")
        print(f"‚è≠Ô∏è  –ü—Ä–æ–ø—É—â–µ–Ω–æ (—É–∂–µ –æ–±—Ä–∞–±–æ—Ç–∞–Ω—ã): {skipped}")
        print(f"üîÑ –û–±—Ä–∞–±–æ—Ç–∞–Ω–æ –∑–∞–Ω–æ–≤–æ: {processed}")
        print(f"‚úÖ –£—Å–ø–µ—à–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∞–Ω–æ: {successful} ({success_rate:.1f}%)")
        print(f"‚ùå –° –æ—à–∏–±–∫–∞–º–∏: {failed}")
        print(f"üîç –í–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏—è —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤:")
        print(f"   ‚úÖ –ü—Ä–æ—à–ª–∏ –ø—Ä–æ–≤–µ—Ä–∫—É: {verification_passed} ({verification_rate:.1f}%)")
        print(f"   ‚ö†Ô∏è  –ü—Ä–æ–±–ª–µ–º—ã –ø—Ä–∏ –ø—Ä–æ–≤–µ—Ä–∫–µ: {verification_failed}")
        
        # –î–µ—Ç–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ –º–µ—Ç–æ–¥–∞–º
        print("-" * 70)
        print("‚öôÔ∏è –î–µ—Ç–∞–ª—å–Ω–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ –º–µ—Ç–æ–¥–∞–º –æ–±—Ä–∞–±–æ—Ç–∫–∏:")
        methods = stats.get('methods', {})
        method_categories = {
            'OCR –º–µ—Ç–æ–¥—ã': ['google_vision', 'tesseract', 'easyocr'],
            '–¢–µ–∫—Å—Ç–æ–≤—ã–µ –º–µ—Ç–æ–¥—ã': ['pypdf_text', 'docx_text', 'xlsx_text'],
            '–°–∏—Å—Ç–µ–º–Ω—ã–µ': ['skipped_existing', 'error', 'unsupported']
        }
        
        for category, method_list in method_categories.items():
            category_total = sum(methods.get(method, 0) for method in method_list)
            if category_total > 0:
                print(f"\n   üìÇ {category}: {category_total} —Ñ–∞–π–ª–æ–≤")
                for method in method_list:
                    count = methods.get(method, 0)
                    if count > 0:
                        percentage = (count / category_total * 100) if category_total > 0 else 0
                        print(f"      ‚Ä¢ {method}: {count} ({percentage:.1f}%)")
        
        # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ —Ç–∏–ø–∞–º —Ñ–∞–π–ª–æ–≤
        file_types = stats.get('file_types', {})
        if file_types:
            print(f"\nüìÅ –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ —Ç–∏–ø–∞–º —Ñ–∞–π–ª–æ–≤:")
            for file_type, type_stats in sorted(file_types.items()):
                type_total = type_stats.get('total', 0)
                type_success = type_stats.get('successful', 0)
                type_rate = (type_success / type_total * 100) if type_total > 0 else 0
                print(f"   ‚Ä¢ {file_type.upper()}: {type_total} —Ñ–∞–π–ª–æ–≤, —É—Å–ø–µ—à–Ω–æ: {type_success} ({type_rate:.1f}%)")
        
        # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å–Ω–æ—Å—Ç–∏
        timing_stats = stats.get('timing', {})
        if timing_stats:
            avg_time = timing_stats.get('average_time', 0)
            total_time = timing_stats.get('total_time', 0)
            fastest = timing_stats.get('fastest', 0)
            slowest = timing_stats.get('slowest', 0)
            print(f"\n‚è±Ô∏è  –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å–Ω–æ—Å—Ç—å:")
            print(f"   ‚Ä¢ –û–±—â–µ–µ –≤—Ä–µ–º—è: {total_time:.1f}—Å")
            print(f"   ‚Ä¢ –°—Ä–µ–¥–Ω–µ–µ –≤—Ä–µ–º—è –Ω–∞ —Ñ–∞–π–ª: {avg_time:.2f}—Å")
            print(f"   ‚Ä¢ –°–∞–º—ã–π –±—ã—Å—Ç—Ä—ã–π: {fastest:.2f}—Å")
            print(f"   ‚Ä¢ –°–∞–º—ã–π –º–µ–¥–ª–µ–Ω–Ω—ã–π: {slowest:.2f}—Å")
        
        # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –∫–∞—á–µ—Å—Ç–≤–∞ OCR
        quality_stats = stats.get('quality', {})
        if quality_stats:
            avg_confidence = quality_stats.get('average_confidence', 0)
            high_confidence = quality_stats.get('high_confidence_count', 0)
            low_confidence = quality_stats.get('low_confidence_count', 0)
            print(f"\nüéØ –ö–∞—á–µ—Å—Ç–≤–æ OCR:")
            print(f"   ‚Ä¢ –°—Ä–µ–¥–Ω—è—è —É–≤–µ—Ä–µ–Ω–Ω–æ—Å—Ç—å: {avg_confidence:.1f}%")
            print(f"   ‚Ä¢ –í—ã—Å–æ–∫–∞—è —É–≤–µ—Ä–µ–Ω–Ω–æ—Å—Ç—å (>80%): {high_confidence} —Ñ–∞–π–ª–æ–≤")
            print(f"   ‚Ä¢ –ù–∏–∑–∫–∞—è —É–≤–µ—Ä–µ–Ω–Ω–æ—Å—Ç—å (<50%): {low_confidence} —Ñ–∞–π–ª–æ–≤")
        
        print("=" * 73)
        
        # –õ–æ–≥–∏—Ä—É–µ–º —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å–Ω–æ—Å—Ç–∏
        self._log_performance_stats(date, stats)
    
    def _log_performance_stats(self, date: str, stats: Dict):
        """üìà –õ–æ–≥–∏—Ä–æ–≤–∞–Ω–∏–µ —Ä–∞—Å—à–∏—Ä–µ–Ω–Ω–æ–π —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å–Ω–æ—Å—Ç–∏"""
        try:
            total = stats.get('total', 0)
            successful = stats.get('successful', 0)
            skipped = stats.get('skipped', 0)
            processed = stats.get('processed', 0)
            failed = processed - successful
            
            verification_passed = stats.get('verification_passed', 0)
            verification_failed = stats.get('verification_failed', 0)
            verification_rate = round((verification_passed / successful * 100) if successful > 0 else 0, 2)
            
            # –ë–∞–∑–æ–≤–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞
            performance_data = {
                 "timestamp": datetime.now().isoformat(),
                 "date": date,
                 "total_files": total,
                 "processed_files": processed,
                 "successful_files": successful,
                 "skipped_files": skipped,
                 "error_files": failed,
                 "verification_passed": verification_passed,
                 "verification_failed": verification_failed,
                 "success_rate": round((successful / processed * 100) if processed > 0 else 0, 2),
                 "error_rate": round((failed / total * 100) if total > 0 else 0, 2),
                 "verification_rate": verification_rate,
                 "methods_used": stats.get('methods', {})
             }
            
            # –î–æ–±–∞–≤–ª—è–µ–º —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É –ø–æ —Ç–∏–ø–∞–º —Ñ–∞–π–ª–æ–≤
            file_types_stats = stats.get('file_types', {})
            if file_types_stats:
                performance_data["file_types_breakdown"] = file_types_stats
            
            # –î–æ–±–∞–≤–ª—è–µ–º —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å–Ω–æ—Å—Ç–∏
            timing_stats = stats.get('timing', {})
            if timing_stats and 'total_time' in timing_stats:
                performance_data["timing_stats"] = {
                    "total_time_sec": round(timing_stats.get('total_time', 0), 2),
                    "average_time_sec": round(timing_stats.get('average_time', 0), 3),
                    "fastest_time_sec": round(timing_stats.get('fastest', 0), 3),
                    "slowest_time_sec": round(timing_stats.get('slowest', 0), 3),
                    "files_per_minute": round((processed / (timing_stats.get('total_time', 1) / 60)) if timing_stats.get('total_time', 0) > 0 else 0, 1)
                }
            
            # –î–æ–±–∞–≤–ª—è–µ–º —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É –∫–∞—á–µ—Å—Ç–≤–∞ OCR
            quality_stats = stats.get('quality', {})
            if quality_stats and 'average_confidence' in quality_stats:
                performance_data["quality_stats"] = {
                    "average_confidence": round(quality_stats.get('average_confidence', 0), 1),
                    "high_confidence_count": quality_stats.get('high_confidence_count', 0),
                    "low_confidence_count": quality_stats.get('low_confidence_count', 0),
                    "confidence_distribution": {
                        "high_confidence_rate": round((quality_stats.get('high_confidence_count', 0) / successful * 100) if successful > 0 else 0, 1),
                        "low_confidence_rate": round((quality_stats.get('low_confidence_count', 0) / successful * 100) if successful > 0 else 0, 1)
                    }
                }
            
            self.logger.info("–†–∞—Å—à–∏—Ä–µ–Ω–Ω–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ –∑–∞–≤–µ—Ä—à–µ–Ω–∞", extra={
                "performance_data": performance_data,
                "event_type": "enhanced_processing_summary"
            })
            
        except Exception as e:
            self.logger.error(f"–û—à–∏–±–∫–∞ –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è —Ä–∞—Å—à–∏—Ä–µ–Ω–Ω–æ–π —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å–Ω–æ—Å—Ç–∏: {e}", exc_info=True)
    
    def _log_operation_time(self, operation_name: str, start_time: float, file_name: str = None, **kwargs):
        """‚è±Ô∏è –õ–æ–≥–∏—Ä–æ–≤–∞–Ω–∏–µ –≤—Ä–µ–º–µ–Ω–∏ –≤—ã–ø–æ–ª–Ω–µ–Ω–∏—è –æ–ø–µ—Ä–∞—Ü–∏–π"""
        try:
            execution_time = time.time() - start_time
            
            log_data = {
                "operation": operation_name,
                "execution_time_sec": round(execution_time, 3),
                "timestamp": datetime.now().isoformat(),
                "event_type": "operation_timing"
            }
            
            if file_name:
                log_data["file_name"] = file_name
            
            # –î–æ–±–∞–≤–ª—è–µ–º –¥–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω—ã–µ –ø–∞—Ä–∞–º–µ—Ç—Ä—ã
            log_data.update(kwargs)
            
            self.logger.debug(f"–û–ø–µ—Ä–∞—Ü–∏—è '{operation_name}' –≤—ã–ø–æ–ª–Ω–µ–Ω–∞ –∑–∞ {execution_time:.3f} —Å–µ–∫", extra=log_data)
            
        except Exception as e:
            self.logger.error(f"–û—à–∏–±–∫–∞ –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è –≤—Ä–µ–º–µ–Ω–∏ –æ–ø–µ—Ä–∞—Ü–∏–∏ {operation_name}: {e}")
    
    def _create_progress_bar(self, current: int, total: int, width: int = 30) -> str:
        """üìä –°–æ–∑–¥–∞–Ω–∏–µ –≤–∏–∑—É–∞–ª—å–Ω–æ–≥–æ –ø—Ä–æ–≥—Ä–µ—Å—Å-–±–∞—Ä–∞"""
        try:
            progress = current / total
            filled = int(width * progress)
            bar = "‚ñà" * filled + "‚ñë" * (width - filled)
            return f"[{bar}]"
        except Exception:
            return "[" + "?" * width + "]"
    
    def _print_current_stats(self, stats: Dict):
        """üìà –û—Ç–æ–±—Ä–∞–∂–µ–Ω–∏–µ —Ç–µ–∫—É—â–µ–π —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏ –æ–±—Ä–∞–±–æ—Ç–∫–∏"""
        try:
            total = stats.get('total', 0)
            successful = stats.get('successful', 0)
            skipped = stats.get('skipped', 0)
            processed = stats.get('processed', 0)
            verification_passed = stats.get('verification_passed', 0)
            verification_failed = stats.get('verification_failed', 0)
            
            success_rate = round((successful / processed * 100) if processed > 0 else 0, 1)
            verification_rate = round((verification_passed / successful * 100) if successful > 0 else 0, 1)
            
            print(f"   üìä –¢–µ–∫—É—â–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞: –í—Å–µ–≥–æ: {total} | –û–±—Ä–∞–±–æ—Ç–∞–Ω–æ: {processed} | –£—Å–ø–µ—à–Ω–æ: {successful} ({success_rate}%) | –ü—Ä–æ–ø—É—â–µ–Ω–æ: {skipped} | –í–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏—è: {verification_passed}/{successful} ({verification_rate}%)")
        except Exception as e:
            print(f"   ‚ö†Ô∏è –û—à–∏–±–∫–∞ –æ—Ç–æ–±—Ä–∞–∂–µ–Ω–∏—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏: {e}")
    def test_files_by_date(self, date: str, files_to_test: List[Path], limit: int = None):
        """üß™ –¢–µ—Å—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ —Ñ–∞–π–ª–æ–≤ –∑–∞ –∫–æ–Ω–∫—Ä–µ—Ç–Ω—É—é –¥–∞—Ç—É —Å –æ–ø—Ç–∏–º–∏–∑–∞—Ü–∏–µ–π –ø–æ–≤—Ç–æ—Ä–Ω–æ–π –æ–±—Ä–∞–±–æ—Ç–∫–∏"""
        if limit:
            files_to_test = files_to_test[:limit]
            print(f"üéØ –û–≥—Ä–∞–Ω–∏—á–µ–Ω–∏–µ: —Ç–µ—Å—Ç–∏—Ä—É–µ–º –ø–µ—Ä–≤—ã–µ {limit} —Ñ–∞–π–ª–æ–≤.")
        
        stats = {
            "total": 0, "successful": 0, "skipped": 0, "processed": 0, 
            "verification_passed": 0, "verification_failed": 0, 
            "methods": {}, "file_types": {}, "timing": {"times": []}, "quality": {"confidences": []}
        }
        total_files = len(files_to_test)
        
        print(f"\nüîç –ü—Ä–æ–≤–µ—Ä—è—é —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã –¥–ª—è {total_files} —Ñ–∞–π–ª–æ–≤...")
        print(f"üìä –ü—Ä–æ–≥—Ä–µ—Å—Å –æ–±—Ä–∞–±–æ—Ç–∫–∏ –¥–ª—è –¥–∞—Ç—ã {date}:")
        print("=" * 80)
        
        for i, file_path in enumerate(files_to_test, 1):
            # –û—Ç–æ–±—Ä–∞–∂–µ–Ω–∏–µ –ø—Ä–æ–≥—Ä–µ—Å—Å–∞
            progress_percent = (i / total_files) * 100
            progress_bar = self._create_progress_bar(i, total_files)
            
            print(f"\n{progress_bar} [{i:3d}/{total_files}] ({progress_percent:5.1f}%)")
            print(f"üìÑ –§–∞–π–ª: {file_path.name}")
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ —É–∂–µ –æ–±—Ä–∞–±–æ—Ç–∞–Ω–Ω—ã–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã
            if self._check_existing_results(file_path, date):
                print(f"   ‚è≠Ô∏è  –°—Ç–∞—Ç—É—Å: –£–ñ–ï –û–ë–†–ê–ë–û–¢–ê–ù - –ø—Ä–æ–ø—É—Å–∫–∞–µ–º")
                stats["total"] += 1
                stats["skipped"] += 1
                stats["methods"]["skipped_existing"] = stats["methods"].get("skipped_existing", 0) + 1
                self._print_current_stats(stats)
                continue
            
            # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Ñ–∞–π–ª
            print(f"   üîÑ –°—Ç–∞—Ç—É—Å: –û–ë–†–ê–ë–ê–¢–´–í–ê–ï–¢–°–Ø...")
            start_time = time.time()
            result = self.extract_text_from_file(file_path, date)
            processing_time = time.time() - start_time
            
            # –°–æ–±–∏—Ä–∞–µ–º —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É –ø–æ —Ç–∏–ø–∞–º —Ñ–∞–π–ª–æ–≤
            file_ext = file_path.suffix.lower().lstrip('.')
            if file_ext not in stats["file_types"]:
                stats["file_types"][file_ext] = {"total": 0, "successful": 0}
            stats["file_types"][file_ext]["total"] += 1
            
            # –°–æ–±–∏—Ä–∞–µ–º —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É –≤—Ä–µ–º–µ–Ω–∏
            stats["timing"]["times"].append(processing_time)
            
            stats["total"] += 1
            stats["processed"] += 1
            
            if result["success"]:
                stats["successful"] += 1
                stats["file_types"][file_ext]["successful"] += 1
                print(f"   ‚úÖ –°—Ç–∞—Ç—É—Å: –£–°–ü–ï–®–ù–û ({result['method']}) –∑–∞ {processing_time:.2f}—Å")
                
                # –°–æ–±–∏—Ä–∞–µ–º —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É –∫–∞—á–µ—Å—Ç–≤–∞ OCR
                confidence = result.get('confidence', 0)
                if confidence > 0:
                    stats["quality"]["confidences"].append(confidence)
                
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ—Å—Ç—å —Å–æ—Ö—Ä–∞–Ω–µ–Ω–Ω—ã—Ö —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤
                if self._verify_saved_result(result, date):
                    stats["verification_passed"] += 1
                    print(f"   üîç –í–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏—è: –ü–†–û–ô–î–ï–ù–ê")
                else:
                    stats["verification_failed"] += 1
                    print(f"   ‚ö†Ô∏è  –í–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏—è: –ü–†–û–ë–õ–ï–ú–´")
            else:
                print(f"   ‚ùå –°—Ç–∞—Ç—É—Å: –û–®–ò–ë–ö–ê ({result.get('error', '–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–∞—è –æ—à–∏–±–∫–∞')})")
            
            stats["methods"][result["method"]] = stats["methods"].get(result["method"], 0) + 1
            self._print_current_stats(stats)
        
        # –í—ã—á–∏—Å–ª—è–µ–º –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞–Ω–Ω—ã–µ —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏
        times = stats["timing"]["times"]
        if times:
            stats["timing"]["total_time"] = sum(times)
            stats["timing"]["average_time"] = sum(times) / len(times)
            stats["timing"]["fastest"] = min(times)
            stats["timing"]["slowest"] = max(times)
        
        confidences = stats["quality"]["confidences"]
        if confidences:
            stats["quality"]["average_confidence"] = sum(confidences) / len(confidences)
            stats["quality"]["high_confidence_count"] = sum(1 for c in confidences if c > 80)
            stats["quality"]["low_confidence_count"] = sum(1 for c in confidences if c < 50)
        
        print("\n" + "=" * 80)
        print(f"üéâ –¢–µ—Å—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –¥–ª—è –¥–∞—Ç—ã {date} –∑–∞–≤–µ—Ä—à–µ–Ω–æ!")
        self._print_summary(date, stats)
def main():
    import argparse
    
    # –ü–∞—Ä—Å–∏–Ω–≥ –∞—Ä–≥—É–º–µ–Ω—Ç–æ–≤ –∫–æ–º–∞–Ω–¥–Ω–æ–π —Å—Ç—Ä–æ–∫–∏
    parser = argparse.ArgumentParser(description='OCR Processor')
    parser.add_argument('--auto', action='store_true', help='–ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏–π —Ä–µ–∂–∏–º –±–µ–∑ –∏–Ω—Ç–µ—Ä–∞–∫—Ç–∏–≤–Ω–æ–≥–æ –º–µ–Ω—é')
    parser.add_argument('--date', type=str, help='–î–∞—Ç–∞ –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏ –≤ —Ñ–æ—Ä–º–∞—Ç–µ YYYY-MM-DD')
    args = parser.parse_args()
    
    try: 
        tester = OCRProcessor()
    except Exception as e:
        print(f"‚ùå –ö—Ä–∏—Ç–∏—á–µ—Å–∫–∞—è –æ—à–∏–±–∫–∞ –ø—Ä–∏ –∏–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏–∏: {e}")
        return
        
    if not tester.vision_client:
        print("\n‚ö†Ô∏è –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –Ω–∞—Å—Ç—Ä–æ–π—Ç–µ Google Cloud Vision –∏ –ø–µ—Ä–µ–∑–∞–ø—É—Å—Ç–∏—Ç–µ —Å–∫—Ä–∏–ø—Ç.")
        return
        
    available_dates = tester.get_available_dates()
    if not available_dates:
        print("ü§∑ –í –ø–∞–ø–∫–µ 'data/attachments' –Ω–µ –Ω–∞–π–¥–µ–Ω–æ –ø–∞–ø–æ–∫ —Å –¥–∞—Ç–∞–º–∏ (YYYY-MM-DD).")
        return
    
    # –ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏–π —Ä–µ–∂–∏–º
    if args.auto:
        if args.date:
            # –û–±—Ä–∞–±–æ—Ç–∫–∞ –∫–æ–Ω–∫—Ä–µ—Ç–Ω–æ–π –¥–∞—Ç—ã
            if args.date not in available_dates:
                print(f"‚ùå –î–∞—Ç–∞ '{args.date}' –Ω–µ –Ω–∞–π–¥–µ–Ω–∞ –≤ –¥–æ—Å—Ç—É–ø–Ω—ã—Ö –¥–∞—Ç–∞—Ö: {', '.join(available_dates)}")
                return
            
            print(f"üöÄ –ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–æ–≤ –∑–∞ {args.date}")
            files_found = tester.get_files_for_date(args.date)
            if not files_found:
                print(f"ü§∑ –í –ø–∞–ø–∫–µ –∑–∞ {args.date} –Ω–µ –Ω–∞–π–¥–µ–Ω–æ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ–º—ã—Ö —Ñ–∞–π–ª–æ–≤.")
                return
            
            print(f"‚úÖ –ù–∞–π–¥–µ–Ω–æ —Ñ–∞–π–ª–æ–≤ –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏: {len(files_found)}")
            tester.test_files_by_date(args.date, files_found)
        else:
            # –û–±—Ä–∞–±–æ—Ç–∫–∞ –≤—Å–µ—Ö –¥–∞—Ç
            print("üöÄ –ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ –≤—Å–µ—Ö —Ñ–∞–π–ª–æ–≤")
            for date in available_dates:
                print(f"\n\n--- üöÄ –û–±—Ä–∞–±–æ—Ç–∫–∞ –¥–∞—Ç—ã: {date} ---")
                files_found = tester.get_files_for_date(date)
                if files_found:
                    tester.test_files_by_date(date, files_found)
                else:
                    print(f"ü§∑ –í –ø–∞–ø–∫–µ –∑–∞ {date} –Ω–µ –Ω–∞–π–¥–µ–Ω–æ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ–º—ã—Ö —Ñ–∞–π–ª–æ–≤.")
        return
    
    # –ò–Ω—Ç–µ—Ä–∞–∫—Ç–∏–≤–Ω—ã–π —Ä–µ–∂–∏–º
    while True:
        print("\n\n" + "="*25 + " üéØ –ú–ï–ù–Æ –¢–ï–°–¢–ò–†–û–í–©–ò–ö–ê üéØ " + "="*25)
        print(f"üìÖ –î–æ—Å—Ç—É–ø–Ω—ã–µ –¥–∞—Ç—ã –¥–ª—è —Ç–µ—Å—Ç–∞: {', '.join(available_dates)}")
        print("1. üß™ –ü—Ä–æ—Ç–µ—Å—Ç–∏—Ä–æ–≤–∞—Ç—å —Ñ–∞–π–ª—ã –∑–∞ –∫–æ–Ω–∫—Ä–µ—Ç–Ω—É—é –¥–∞—Ç—É")
        print("2. üöÄ –ü—Ä–æ—Ç–µ—Å—Ç–∏—Ä–æ–≤–∞—Ç—å –í–°–ï —Ñ–∞–π–ª—ã –∏–∑ –í–°–ï–• –¥–∞—Ç")
        print("3. üö™ –í—ã–π—Ç–∏")
        choice = input("üëâ –í–∞—à –≤—ã–±–æ—Ä (1-3): ").strip()
        if choice == "1":
            date = input("   –í–≤–µ–¥–∏—Ç–µ –¥–∞—Ç—É (YYYY-MM-DD): ").strip()
            if date not in available_dates:
                print(f"   ‚ùå –î–∞—Ç–∞ '{date}' –Ω–µ –Ω–∞–π–¥–µ–Ω–∞!"); continue
            files_found = tester.get_files_for_date(date)
            if not files_found:
                print(f"ü§∑ –í –ø–∞–ø–∫–µ –∑–∞ {date} –Ω–µ –Ω–∞–π–¥–µ–Ω–æ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ–º—ã—Ö —Ñ–∞–π–ª–æ–≤.")
                continue
            print(f"‚úÖ –ù–∞–π–¥–µ–Ω–æ —Ñ–∞–π–ª–æ–≤ –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏: {len(files_found)}")
            limit_input = input(f"   –°–∫–æ–ª—å–∫–æ —Ñ–∞–π–ª–æ–≤ —Ç–µ—Å—Ç–∏—Ä–æ–≤–∞—Ç—å? (Enter = –≤—Å–µ {len(files_found)}): ").strip()
            limit = int(limit_input) if limit_input.isdigit() else None
            tester.test_files_by_date(date, files_found, limit)
        elif choice == "2":
            if input("   –í—ã —É–≤–µ—Ä–µ–Ω—ã, —á—Ç–æ —Ö–æ—Ç–∏—Ç–µ –ø—Ä–æ—Ç–µ—Å—Ç–∏—Ä–æ–≤–∞—Ç—å –≤—Å–µ —Ñ–∞–π–ª—ã? (y/n): ").lower() == 'y':
                for date in available_dates:
                    print(f"\n\n--- üöÄ –û–±—Ä–∞–±–æ—Ç–∫–∞ –¥–∞—Ç—ã: {date} ---")
                    files_found = tester.get_files_for_date(date)
                    if files_found:
                        tester.test_files_by_date(date, files_found)
                    else:
                        print(f"ü§∑ –í –ø–∞–ø–∫–µ –∑–∞ {date} –Ω–µ –Ω–∞–π–¥–µ–Ω–æ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ–º—ã—Ö —Ñ–∞–π–ª–æ–≤.")
            else:
                print("   –û—Ç–º–µ–Ω–µ–Ω–æ.")
        elif choice == "3":
            print("üëã –î–æ —Å–≤–∏–¥–∞–Ω–∏—è!"); break
        else:
            print("   ‚ùå –ù–µ–≤–µ—Ä–Ω—ã–π –≤—ã–±–æ—Ä. –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤–≤–µ–¥–∏—Ç–µ —á–∏—Å–ª–æ –æ—Ç 1 –¥–æ 3.")

if __name__ == "__main__":
    main()