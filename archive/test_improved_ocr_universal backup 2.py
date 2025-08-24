#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🎯 Гибридный OCR тестер v12 (Google Cloud Vision Edition)
- Отключена защита от "Decompression Bomb" в Pillow для обработки очень больших изображений.
"""

import json
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple
import time
from datetime import datetime
import subprocess
import shutil
import io

from PIL import Image
from google.api_core import exceptions as google_exceptions

# <<< ИЗМЕНЕНИЕ: Отключаем лимит на размер изображения в Pillow >>>
Image.MAX_IMAGE_PIXELS = None

# ... (остальные импорты без изменений) ...
try:
    from google.cloud import vision
    GOOGLE_VISION_AVAILABLE = True
except ImportError:
    GOOGLE_VISION_AVAILABLE = False
try:
    import fitz
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False
try:
    from docx import Document as DocxDocument
    PYTHON_DOCX_AVAILABLE = True
except ImportError:
    PYTHON_DOCX_AVAILABLE = False
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

class GoogleVisionOCRTester:
    # ... (весь остальной код класса остается АБСОЛЮТНО БЕЗ ИЗМЕНЕНИЙ) ...
    def __init__(self):
        self.data_dir = Path("data")
        self.attachments_dir = self.data_dir / "attachments"
        self.base_results_dir = self.data_dir / "final_results"
        self.texts_dir = self.base_results_dir / "texts"
        self.reports_dir = self.base_results_dir / "reports"
        self.texts_dir.mkdir(parents=True, exist_ok=True)
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        self.vision_client = vision.ImageAnnotatorClient() if GOOGLE_VISION_AVAILABLE else None
        self._show_capabilities()
        print("\n" + "=" * 70)
        print("🎯 OCR ТЕСТЕР С GOOGLE CLOUD VISION v12 🎯")
        print(f"📁 Исходные файлы: {self.attachments_dir}")
        print(f"🗂️  Результаты в папке: {self.base_results_dir}")
        print("=" * 70)
    def _show_capabilities(self):
        antiword_ok = shutil.which('antiword') is not None
        print("📋 ВОЗМОЖНОСТИ СИСТЕМЫ:")
        if GOOGLE_VISION_AVAILABLE and self.vision_client:
            print("   ☁️ Google Cloud Vision: ✅ Готов к работе!")
        else:
            print("   ☁️ Google Cloud Vision: ❌ НЕ НАСТРОЕН!")
        local_status = [f"PDF (текст) {'✅' if PYMUPDF_AVAILABLE else '❌'}", f"DOCX {'✅' if PYTHON_DOCX_AVAILABLE else '❌'}", f"XLSX {'✅' if OPENPYXL_AVAILABLE else '❌'}", f"DOC (antiword) {'✅' if antiword_ok else '❌ (brew install antiword)'}", f"XLS (xlrd) {'✅' if XLRD_AVAILABLE else '❌'}"]
        print(f"   📄 Локальные форматы: {' | '.join(local_status)}")
    def get_available_dates(self) -> List[str]:
        if not self.attachments_dir.exists(): return []
        return sorted([d.name for d in self.attachments_dir.iterdir() if d.is_dir() and d.name.count("-") == 2])
    def get_files_for_date(self, date: str) -> List[Path]:
        date_dir = self.attachments_dir / date
        if not date_dir.exists():
            print(f"❌ Папка не существует: {date_dir}")
            return []
        file_types = ["*.png", "*.jpg", "*.jpeg", "*.tiff", "*.pdf", "*.docx", "*.doc", "*.xlsx", "*.xls"]
        files = sorted(list(set(f for pat in file_types for f in date_dir.rglob(pat))))
        return files
    def run_google_vision_ocr(self, content: bytes) -> Tuple[str, float]:
        if not self.vision_client: raise RuntimeError("Клиент Google Vision не инициализирован.")
        print("   ☁️ Отправка в Google Cloud Vision... (может занять несколько секунд)")
        ts = time.time()
        image = vision.Image(content=content)
        response = self.vision_client.document_text_detection(image=image)
        if response.error.message:
            raise Exception(f"Google Vision API Error: {response.error.message}")
        text = response.full_text_annotation.text
        confidences = [page.confidence for page in response.full_text_annotation.pages]
        avg_confidence = np.mean(confidences) if confidences else 0.0
        elapsed = time.time() - ts
        print(f"   ✨ Получен ответ от Google за {elapsed:.2f} сек. Уверенность: {avg_confidence:.2%}")
        return text, avg_confidence
    def extract_text_from_file(self, file_path: Path) -> Dict:
        ext = file_path.suffix.lower()
        text, method, confidence, error = "", "unknown", 0.0, None
        ts = time.time()
        try:
            if ext == ".docx":
                print("   📄 Обработка DOCX локально...")
                doc = DocxDocument(file_path)
                text = "\n".join([p.text for p in doc.paragraphs])
                method, confidence = "local_docx", 1.0
            elif ext == ".doc":
                print("   📄 Обработка DOC (старый формат) через antiword...")
                if not shutil.which('antiword'):
                    raise FileNotFoundError("Утилита 'antiword' не найдена. Установите ее: brew install antiword")
                process = subprocess.run(['antiword', str(file_path)], capture_output=True, text=True, encoding='utf-8', errors='ignore')
                if process.returncode == 0:
                    text = process.stdout
                else:
                    raise RuntimeError(f"Antiword вернул ошибку: {process.stderr}")
                method, confidence = "local_doc_antiword", 1.0
            elif ext == ".xlsx":
                print("   📄 Обработка XLSX локально...")
                wb = openpyxl.load_workbook(file_path, data_only=True)
                lines = [" | ".join([str(cell.value or "") for cell in row]) for sheet in wb.worksheets for row in sheet.iter_rows()]
                text = "\n".join(lines)
                method, confidence = "local_xlsx", 1.0
            elif ext == ".xls":
                print("   📄 Обработка XLS (старый формат) локально...")
                if not XLRD_AVAILABLE: raise ImportError("Библиотека xlrd не найдена. Установите: pip install xlrd")
                wb = xlrd.open_workbook(file_path, encoding_override="cp1251")
                lines = []
                for sheet in wb.sheets():
                    for row_idx in range(sheet.nrows):
                        lines.append(" | ".join([str(sheet.cell(row_idx, col_idx).value or "") for col_idx in range(sheet.ncols)]))
                text = "\n".join(lines)
                method, confidence = "local_xls", 1.0
            elif ext == ".pdf":
                print("   📄 Обработка PDF... Попытка извлечь текстовый слой.")
                doc = fitz.open(file_path)
                texts = [page.get_text() for page in doc]
                full_text_direct = "\n\n".join(texts).strip()
                if len(full_text_direct) > 100:
                    print("   ✅ Обнаружен текстовый слой. Извлечено локально.")
                    text, method, confidence = full_text_direct, "local_pdf_text", 1.0
                else:
                    print("   🖼️ Текстовый слой пуст. Конвертируем страницы PDF в картинки для Google Vision.")
                    all_pages_text = []
                    all_confidences = []
                    for i, page in enumerate(doc):
                        print(f"     -- Обработка страницы {i+1}/{len(doc)} --")
                        pix = page.get_pixmap(dpi=300)
                        try:
                            img_bytes = pix.tobytes("png")
                            page_text, page_confidence = self.run_google_vision_ocr(img_bytes)
                        except google_exceptions.InvalidArgument as e:
                            if "Request payload size exceeds the limit" in str(e):
                                print("     ⚠️ Размер страницы слишком велик! Пробую сжать и нормализовать через Pillow...")
                                png_buffer = io.BytesIO(pix.tobytes("png"))
                                pil_image = Image.open(png_buffer)
                                if pil_image.mode != 'RGB':
                                    pil_image = pil_image.convert('RGB')
                                jpeg_buffer = io.BytesIO()
                                pil_image.save(jpeg_buffer, format="JPEG", quality=92)
                                img_bytes_jpeg = jpeg_buffer.getvalue()
                                page_text, page_confidence = self.run_google_vision_ocr(img_bytes_jpeg)
                            else:
                                raise
                        all_pages_text.append(page_text)
                        all_confidences.append(page_confidence)
                    text = "\n\n--- PAGE BREAK ---\n\n".join(all_pages_text)
                    confidence = np.mean(all_confidences) if all_confidences else 0.0
                    method = "google_vision_pdf"
            elif ext in [".png", ".jpg", ".jpeg", ".tiff"]:
                print(f"   🖼️ Обработка изображения ({ext}). Отправляем в Google Vision.")
                try:
                    img_bytes = file_path.read_bytes()
                    text, confidence = self.run_google_vision_ocr(img_bytes)
                except google_exceptions.InvalidArgument as e:
                    if "Request payload size exceeds the limit" in str(e):
                        print("     ⚠️ Изображение слишком велико! Пробую сжать и нормализовать через Pillow...")
                        with Image.open(file_path) as pil_image:
                            if pil_image.mode != 'RGB':
                                pil_image = pil_image.convert('RGB')
                            img_buffer = io.BytesIO()
                            pil_image.save(img_buffer, format="JPEG", quality=92)
                            img_bytes_jpeg = img_buffer.getvalue()
                        text, confidence = self.run_google_vision_ocr(img_bytes_jpeg)
                    else:
                        raise
                method = "google_vision_image"
            else:
                method, error = "unsupported", f"Формат {ext} не поддерживается."
        except Exception as e:
            error = str(e)
            method = "error"
            print(f"   ❌ Произошла ошибка: {e}")
        return {"file_path": str(file_path), "file_name": file_path.name, "success": not error and bool(text.strip()), "text": text.strip(), "method": method, "confidence": confidence, "error": error, "processing_time_sec": time.time() - ts, "timestamp": datetime.now().isoformat()}
    def save_result(self, result: Dict, date: str):
        date_texts_dir = self.texts_dir / date
        date_texts_dir.mkdir(parents=True, exist_ok=True)
        if result["success"]:
            txt_filename = f"{Path(result['file_name']).stem}___{result['method']}.txt"
            txt_path = date_texts_dir / txt_filename
            with open(txt_path, "w", encoding="utf-8") as f:
                f.write(f"# 📄 Файл: {result['file_name']}\n# ⚙️ Метод: {result['method']}\n")
                f.write(f"# ✨ Уверенность: {result['confidence']:.2%}\n# ⏱️ Время: {result['processing_time_sec']:.2f} сек\n")
                f.write("# " + "=" * 50 + "\n\n" + result['text'])
            print(f"   💾 TXT результат сохранен в: {txt_path}")
            result["txt_file_path"] = str(txt_path)
        json_report_path = self.reports_dir / f"test_report_{date}.json"
        report_data = []
        if json_report_path.exists():
            with open(json_report_path, "r", encoding="utf-8") as f:
                try: report_data = json.load(f)
                except json.JSONDecodeError: pass
        report_data.append(result)
        with open(json_report_path, "w", encoding="utf-8") as f:
            json.dump(report_data, f, ensure_ascii=False, indent=2)
    def _print_summary(self, date: str, stats: Dict):
        print("\n" + "="*25 + f" 📊 ИТОГИ ТЕСТА ЗА {date} " + "="*25)
        total = stats.get('total', 0)
        successful = stats.get('successful', 0)
        failed = total - successful
        success_rate = (successful / total * 100) if total > 0 else 0
        print(f"🗂️  Всего обработано файлов: {total}")
        print(f"✅ Успешно: {successful} ({success_rate:.1f}%)")
        print(f"❌ С ошибками: {failed}")
        print("-" * 70)
        print("⚙️ Статистика по методам:")
        methods = stats.get('methods', {})
        for method, count in sorted(methods.items()):
            print(f"   - {method}: {count} раз")
        print("=" * 73)
    def test_files_by_date(self, date: str, files_to_test: List[Path], limit: int = None):
        if limit:
            files_to_test = files_to_test[:limit]
            print(f"🎯 Ограничение: тестируем первые {limit} файлов.")
        stats = {"total": 0, "successful": 0, "methods": {}}
        for i, file_path in enumerate(files_to_test, 1):
            print(f"\n--- 🧪 Файл {i}/{len(files_to_test)}: {file_path.name} ---")
            result = self.extract_text_from_file(file_path)
            self.save_result(result, date)
            stats["total"] += 1
            if result["success"]:
                stats["successful"] += 1
            stats["methods"][result["method"]] = stats["methods"].get(result["method"], 0) + 1
        print(f"\n🎉 Тестирование для даты {date} завершено!")
        self._print_summary(date, stats)
def main():
    try: tester = GoogleVisionOCRTester()
    except Exception as e:
        print(f"❌ Критическая ошибка при инициализации: {e}"); return
    if not tester.vision_client:
        print("\n⚠️ Пожалуйста, настройте Google Cloud Vision и перезапустите скрипт."); return
    available_dates = tester.get_available_dates()
    if not available_dates:
        print("🤷 В папке 'data/attachments' не найдено папок с датами (YYYY-MM-DD)."); return
    while True:
        print("\n\n" + "="*25 + " 🎯 МЕНЮ ТЕСТИРОВЩИКА 🎯 " + "="*25)
        print(f"📅 Доступные даты для теста: {', '.join(available_dates)}")
        print("1. 🧪 Протестировать файлы за конкретную дату")
        print("2. 🚀 Протестировать ВСЕ файлы из ВСЕХ дат")
        print("3. 🚪 Выйти")
        choice = input("👉 Ваш выбор (1-3): ").strip()
        if choice == "1":
            date = input("   Введите дату (YYYY-MM-DD): ").strip()
            if date not in available_dates:
                print(f"   ❌ Дата '{date}' не найдена!"); continue
            files_found = tester.get_files_for_date(date)
            if not files_found:
                print(f"🤷 В папке за {date} не найдено поддерживаемых файлов.")
                continue
            print(f"✅ Найдено файлов для обработки: {len(files_found)}")
            limit_input = input(f"   Сколько файлов тестировать? (Enter = все {len(files_found)}): ").strip()
            limit = int(limit_input) if limit_input.isdigit() else None
            tester.test_files_by_date(date, files_found, limit)
        elif choice == "2":
            if input("   Вы уверены, что хотите протестировать все файлы? (y/n): ").lower() == 'y':
                for date in available_dates:
                    print(f"\n\n--- 🚀 Обработка даты: {date} ---")
                    files_found = tester.get_files_for_date(date)
                    if files_found:
                        tester.test_files_by_date(date, files_found)
                    else:
                        print(f"🤷 В папке за {date} не найдено поддерживаемых файлов.")
            else:
                print("   Отменено.")
        elif choice == "3":
            print("👋 До свидания!"); break
        else:
            print("   ❌ Неверный выбор. Пожалуйста, введите число от 1 до 3.")

if __name__ == "__main__":
    main()